import streamlit as st
import subprocess
import sys
import os
import re
import time
import tempfile
from io import BytesIO
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage

# ───────────────────────────────────────────────
# Install Playwright browsers on first run
# ───────────────────────────────────────────────
@st.cache_resource
def install_playwright():
    subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        check=True, capture_output=True,
    )
    return True

# ───────────────────────────────────────────────
# Page Config
# ───────────────────────────────────────────────
st.set_page_config(page_title="L&D Page Extractor", page_icon="🔍", layout="wide")
st.markdown("""
<style>
    .stApp { background-color: #f8f9fb; }
    h1 { font-size: 1.8rem !important; font-weight: 700; color: #1a1a2e; }
</style>
""", unsafe_allow_html=True)
st.title("🔍 L&D Competitor Page Extractor")
st.caption("Uses Playwright (Chromium) to screenshot each page, then extracts only the main body content.")

DEFAULT_URLS = """https://cmoe.com/products-and-services/learning-and-development-advisory-services/
https://www.ey.com/en_gl/services/workforce/learning-development-advisory
https://www.optimuslearningservices.com/l-and-d-services/consult/
https://www.eidesign.net/ld-consulting/
https://wdhb.com/services/strategy-design-consulting/
https://www.futuristsspeakers.com/learning-development-consulting-services/
https://www.hemsleyfraser.com/en-us/consultancy-services
https://www.thinkdom.co/learning-and-development-consulting
https://www.wipro.com/consulting/learning-and-development-consulting-services/
https://services.elblearning.com/learning-and-development-consulting"""

# ───────────────────────────────────────────────
# Null-safe JS extraction — runs inside the page
# ───────────────────────────────────────────────
EXTRACT_JS = """
() => {
    try {
        // Step 1: Remove noise elements
        const noiseSelectors = [
            'header', 'nav', 'footer',
            '[role="navigation"]', '[role="banner"]', '[role="contentinfo"]',
            '.cookie-banner', '.cookie-consent', '#onetrust-banner-sdk',
            '.site-header', '.site-footer', '.site-nav',
            '.mega-menu', '.mobile-menu', '.breadcrumb', '.breadcrumbs',
            '.social-share', '.share-buttons',
            '.sidebar', 'aside',
            'script', 'style', 'noscript', 'iframe',
            'svg', 'picture', 'video', 'canvas'
        ];

        noiseSelectors.forEach(function(sel) {
            try {
                var els = document.querySelectorAll(sel);
                for (var i = 0; i < els.length; i++) {
                    try { els[i].parentNode.removeChild(els[i]); } catch(e) {}
                }
            } catch(e) {}
        });

        // Step 2: Remove by class/id patterns
        var noisePattern = /nav|header|footer|menu|sidebar|cookie|consent|banner|popup|modal|breadcrumb|social|share|skip|onetrust|grecaptcha/i;
        var allEls = document.querySelectorAll('div, section, aside, ul, ol, span, a');
        for (var i = 0; i < allEls.length; i++) {
            var el = allEls[i];
            try {
                var cls = el.className || '';
                if (typeof cls !== 'string') cls = '';
                var elId = el.id || '';
                var role = el.getAttribute('role') || '';
                if (noisePattern.test(cls) || noisePattern.test(elId) ||
                    role === 'navigation' || role === 'banner' || role === 'contentinfo') {
                    if (el.parentNode) el.parentNode.removeChild(el);
                }
            } catch(e) {}
        }

        // Step 3: Remove images
        var imgs = document.querySelectorAll('img');
        for (var i = 0; i < imgs.length; i++) {
            try { if (imgs[i].parentNode) imgs[i].parentNode.removeChild(imgs[i]); } catch(e) {}
        }

        // Step 4: Find main content
        var main = document.querySelector('main')
            || document.querySelector('article')
            || document.querySelector('[role="main"]')
            || document.querySelector('#content')
            || document.querySelector('#main-content')
            || document.querySelector('.main-content')
            || document.querySelector('.page-content')
            || document.querySelector('.entry-content')
            || document.querySelector('.content')
            || document.body;

        if (!main) return '[No content container found]';

        // Step 5: Extract text with structure (null-safe)
        function getText(node) {
            if (!node) return '';
            var result = '';
            var children = node.childNodes;
            if (!children) return node.textContent || '';

            for (var i = 0; i < children.length; i++) {
                var child = children[i];
                if (!child) continue;

                // Text node
                if (child.nodeType === 3) {
                    var txt = (child.textContent || '').trim();
                    if (txt) result += txt + ' ';
                    continue;
                }

                // Element node
                if (child.nodeType !== 1) continue;

                var tag = (child.tagName || '').toLowerCase();
                if (!tag) continue;

                // Skip invisible elements
                try {
                    var style = window.getComputedStyle(child);
                    if (style && (style.display === 'none' || style.visibility === 'hidden')) continue;
                } catch(e) {}

                var blockTags = {div:1, p:1, h1:1, h2:1, h3:1, h4:1, h5:1, h6:1,
                                 li:1, tr:1, section:1, article:1, blockquote:1, figcaption:1, dt:1, dd:1};

                if (blockTags[tag]) result += '\\n';

                // Heading markers
                if (tag === 'h1') result += '\\n# ';
                else if (tag === 'h2') result += '\\n## ';
                else if (tag === 'h3') result += '\\n### ';
                else if (tag === 'h4') result += '\\n#### ';
                else if (tag === 'h5') result += '\\n##### ';
                else if (tag === 'h6') result += '\\n###### ';

                // List markers
                if (tag === 'li') result += '• ';

                // Line break
                if (tag === 'br') { result += '\\n'; continue; }

                result += getText(child);

                if (blockTags[tag]) result += '\\n';
            }
            return result;
        }

        var text = getText(main);
        // Clean whitespace
        text = text.replace(/[ \\t]+/g, ' ');
        text = text.replace(/\\n[ \\t]+/g, '\\n');
        text = text.replace(/\\n{3,}/g, '\\n\\n');
        return text.trim();

    } catch(err) {
        return '[JS Error: ' + err.message + ']';
    }
}
"""


# ───────────────────────────────────────────────
# Core extraction
# ───────────────────────────────────────────────
def extract_page(url, timeout_ms, screenshot_dir):
    from playwright.sync_api import sync_playwright

    domain = get_domain(url)
    screenshot_path = os.path.join(screenshot_dir, f"{domain}.png")

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"]
            )
            context = browser.new_context(
                viewport={"width": 1440, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/122.0.0.0 Safari/537.36"
                ),
            )
            page = context.new_page()
            page.goto(url, wait_until="networkidle", timeout=timeout_ms)
            page.wait_for_timeout(2000)

            # Scroll to trigger lazy loading
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1500)
            page.evaluate("window.scrollTo(0, 0)")
            page.wait_for_timeout(500)

            # Screenshot BEFORE removing elements
            page.screenshot(path=screenshot_path, full_page=True, type="png")

            # Extract content via JS
            content = page.evaluate(EXTRACT_JS)
            browser.close()

            if not content or len(content.strip()) < 100:
                return {
                    "url": url, "status": "warning",
                    "content": content or "[Minimal content extracted]",
                    "screenshot": screenshot_path if os.path.exists(screenshot_path) else None,
                    "method": "playwright-minimal",
                }

            return {
                "url": url, "status": "success",
                "content": content.strip(),
                "screenshot": screenshot_path,
                "method": "playwright",
            }

    except Exception as e:
        return {
            "url": url, "status": "error",
            "content": f"Error: {str(e)[:300]}",
            "screenshot": screenshot_path if os.path.exists(screenshot_path) else None,
            "method": "failed",
        }


# ───────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────
def get_domain(url):
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.replace("www.", "").replace("services.", "")
        return re.sub(r"[^a-zA-Z0-9]", "_", domain.split(".")[0].capitalize())
    except Exception:
        return "unknown"


def build_excel(results, include_screenshots=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Page Body Content"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    body_font = Font(name="Arial", size=9)
    wrap = Alignment(wrap_text=True, vertical="top")
    center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
    border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
    )

    if include_screenshots:
        headers = ["S.No", "Company", "URL", "Method", "Screenshot", "Body Content"]
        widths = {"A": 6, "B": 18, "C": 50, "D": 14, "E": 40, "F": 140}
    else:
        headers = ["S.No", "Company", "URL", "Method", "Body Content"]
        widths = {"A": 6, "B": 18, "C": 50, "D": 14, "E": 160}

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center_wrap
        c.border = border

    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for idx, r in enumerate(results, 1):
        row = idx + 1
        domain = get_domain(r["url"])
        content = r["content"][:32000] if r["content"] else ""
        method = r.get("method", "?")

        if include_screenshots:
            values = [idx, domain, r["url"], method, "", content]
        else:
            values = [idx, domain, r["url"], method, content]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = body_font
            c.alignment = wrap
            c.border = border

        if include_screenshots and r.get("screenshot") and os.path.exists(r["screenshot"]):
            try:
                img = PILImage.open(r["screenshot"])
                img.thumbnail((300, 600), PILImage.LANCZOS)
                thumb_path = r["screenshot"].replace(".png", "_thumb.png")
                img.save(thumb_path)
                xl_img = XlImage(thumb_path)
                ws.add_image(xl_img, f"E{row}")
                ws.row_dimensions[row].height = max(300, int(img.height * 0.75))
            except Exception:
                ws.row_dimensions[row].height = 400
        else:
            ws.row_dimensions[row].height = 400

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ───────────────────────────────────────────────
# Sidebar
# ───────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")
    timeout_sec = st.slider("Page load timeout (sec)", 10, 60, 30)
    include_screenshots = st.checkbox("Include screenshots in Excel", value=True)
    st.markdown("---")
    st.subheader("How it works")
    st.markdown(
        "**1.** Launches **Chromium** (headless) via Playwright\n\n"
        "**2.** Loads each URL, waits for JS, scrolls to trigger lazy loading\n\n"
        "**3.** Takes a **full-page screenshot** before DOM manipulation\n\n"
        "**4.** Runs JS to **remove** nav, header, footer, sidebar, cookies\n\n"
        "**5.** Extracts **main content** with headings and list structure\n\n"
        "**6.** Exports to Excel with screenshots + clean text"
    )


# ───────────────────────────────────────────────
# Main UI
# ───────────────────────────────────────────────
with st.spinner("🔧 Setting up browser engine (first run only)..."):
    install_playwright()

urls_input = st.text_area("📋 Paste URLs (one per line):", value=DEFAULT_URLS, height=260)

col1, col2, _ = st.columns([1, 1, 3])
with col1:
    run_btn = st.button("🚀 Extract All", type="primary", use_container_width=True)
with col2:
    if st.button("🗑️ Clear Results", use_container_width=True):
        st.session_state.pop("results", None)
        st.rerun()

if run_btn:
    urls = [u.strip() for u in urls_input.strip().split("\n") if u.strip()]
    if not urls:
        st.error("Please enter at least one URL.")
    else:
        screenshot_dir = tempfile.mkdtemp(prefix="ld_screenshots_")
        results = []
        timeout_ms = timeout_sec * 1000
        progress = st.progress(0, text="Starting browser...")

        for i, url in enumerate(urls):
            domain = get_domain(url)
            progress.progress(i / len(urls), text=f"🌐 Loading {i+1}/{len(urls)}: {domain}...")
            result = extract_page(url, timeout_ms, screenshot_dir)
            results.append(result)

        progress.progress(1.0, text="✅ All done!")
        st.session_state["results"] = results
        st.session_state["screenshot_dir"] = screenshot_dir


# ───────────────────────────────────────────────
# Results
# ───────────────────────────────────────────────
if "results" in st.session_state:
    results = st.session_state["results"]
    success = sum(1 for r in results if r["status"] == "success")
    warnings = sum(1 for r in results if r["status"] == "warning")
    errors = sum(1 for r in results if r["status"] == "error")

    st.markdown("---")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total", len(results))
    c2.metric("✅ Extracted", success)
    c3.metric("⚠️ Partial", warnings)
    c4.metric("❌ Failed", errors)

    excel_buf = build_excel(results, include_screenshots)
    st.download_button(
        label="📥 Download Excel",
        data=excel_buf,
        file_name="LD_Competitor_Content.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    st.markdown("---")
    st.subheader("📄 Extracted Content Preview")

    for r in results:
        domain = get_domain(r["url"])
        content_len = len(r["content"]) if r["content"] else 0
        method = r.get("method", "?")
        icon = "✅" if r["status"] == "success" else ("⚠️" if r["status"] == "warning" else "❌")

        with st.expander(f"{icon} {domain} — {content_len:,} chars (via {method})", expanded=False):
            st.caption(r["url"])
            if r.get("screenshot") and os.path.exists(r["screenshot"]):
                col_img, col_txt = st.columns([1, 2])
                with col_img:
                    st.image(r["screenshot"], caption="Screenshot", use_container_width=True)
                with col_txt:
                    if r["status"] == "error":
                        st.error(r["content"])
                    else:
                        preview = r["content"][:5000]
                        if content_len > 5000:
                            preview += "\n\n... [full content in Excel]"
                        st.text(preview)
            else:
                if r["status"] == "error":
                    st.error(r["content"])
                else:
                    st.text(r["content"][:5000])
